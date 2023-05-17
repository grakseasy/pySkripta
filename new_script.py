from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re

filename = 'ahb_coverage.sv'
svfile = open(filename, "w")

wb = Workbook()

path_file = 'covergroup_ahb.xlsx'
wb = load_workbook(path_file)
ws = wb.active

for row in ws.iter_rows(min_row=2, max_col=7, max_row=2):
    for cell in row:
        print(cell.value)

coverageGroup = ws.cell(2,1).value
coverpointName = ws.cell(2, 3).value

#GOTTA LOOP THROUGH THAT SHIT
with open('ahb_coverage.sv', 'w') as f:
    f.write('covergroup ' + coverageGroup + ';\n')
    f.write('\n')
    f.write('  coverpoint_' + coverpointName + ' {\n')
    f.write('    bins' + coverpointName + ' = { bin_expr };\n')
    f.write('  }\n')
    f.write('\n')
    f.write('endgroup\n')

os.startfile(filename)
