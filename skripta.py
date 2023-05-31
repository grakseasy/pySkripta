#Define the covergroup name
from openpyxl import load_workbook
import re

wb = load_workbook('covergroup_ahb.xlsx')
ws = wb['Sheet2'] 

column_data = []
for row in ws.iter_rows(min_row=2, min_col=1, max_col=15, values_only=True):
    for cell_value in row:
        if cell_value is not None:
            column_data.append(cell_value)

covergroup_name = column_data.pop(0)
column_data.pop(0) #This leaves me with only SPLIT_BINS columns

signals = []
for signal in column_data:
    string = signal.split(":",1)[1]
    match = re.search(r'(\w+)=', string)
    result = match.group(1)
    signals.append(result) #'HADDR', 'HBURST'...

coverpoints = list(zip(signals, column_data))

with open('ahb_coverage.sv', 'w') as file:
    file.write('`define cov\n')
    file.write('`ifdef COVERAGE_EN\n\n')
    file.write('covergroup ' + covergroup_name + ' with function sample(sequence_item item);\n')
    file.write('covergroup ' + covergroup_name + ';\n')
    file.write('option.per_instance = ' + str(1) + ';\n')
    file.write('option.name = "' + covergroup_name + '";\n\n')
    for cp_name, cp_bins in coverpoints:
        file.write('    coverpoint_' + cp_name + '      : coverpoint item.' + cp_name + '\n')
        file.write('    {\n')
        for i, bin_range in enumerate(cp_bins.split('{')[1].split('}')[0].split(',')):
            file.write('        bins    ' + cp_name + '_' + str(i) + ' = {' + bin_range.strip() + '};\n')
        file.write('    }\n')
    file.write('endgroup\n')
    file.write('\n\n`endif')

#coverpoints list: ('HRDATA', 'SPLIT_BINS:HRDATA= { [32'h0:32'h66666665], 32'h66666666:32'hcccccccb, 32'hcccccccc:32'hffffffff}'), ...
